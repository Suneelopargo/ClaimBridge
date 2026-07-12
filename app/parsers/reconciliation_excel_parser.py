from __future__ import annotations

import hashlib
import json
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.parsers.reconciliation_field_mapping import (
    DATE_FIELDS,
    DECIMAL_FIELDS,
    EXPECTED_IHX_HEADERS,
    IHX_RECONCILIATION_FIELD_MAPPING,
)


logger = logging.getLogger(__name__)


class ReconciliationParserError(Exception):
    """Raised when the reconciliation workbook cannot be parsed safely."""


@dataclass(frozen=True)
class ReconciliationRow:
    source_row_number: int
    values: Dict[str, Any]
    change_hash: str

    @property
    def ihx_ref_id(self) -> str:
        return self.values["ihx_ref_id"]


@dataclass(frozen=True)
class ReconciliationRowError:
    source_row_number: int
    message: str


@dataclass
class ReconciliationParseResult:
    source_file_name: str
    sheet_name: str
    total_excel_rows: int = 0
    parsed_rows: List[ReconciliationRow] = field(default_factory=list)
    skipped_blank_rows: int = 0
    row_errors: List[ReconciliationRowError] = field(default_factory=list)

    @property
    def successful_rows(self) -> int:
        return len(self.parsed_rows)

    @property
    def failed_rows(self) -> int:
        return len(self.row_errors)


class ReconciliationExcelParser:
    """
    Parses the complete IHX Power BI reconciliation export.

    Responsibilities:
    - validate the exact 43-column export structure;
    - normalize strings, dates and decimal values;
    - reject rows without IHX Ref Id;
    - calculate a stable hash for change detection;
    - remain independent from SQLAlchemy and database code.
    """

    def parse(self, file_path: str | Path) -> ReconciliationParseResult:
        path = Path(file_path)

        if not path.exists():
            raise ReconciliationParserError(
                f"Reconciliation workbook does not exist: {path}"
            )

        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ReconciliationParserError(
                f"Unsupported reconciliation workbook type: {path.suffix}"
            )

        logger.info(
            "[ReconciliationParser] Parsing workbook: %s",
            path,
        )

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ReconciliationParserError(
                f"Unable to open reconciliation workbook: {path.name}"
            ) from exc

        try:
            worksheet = self._select_worksheet(workbook)
            headers = self._read_headers(worksheet)
            self._validate_headers(headers)

            result = ReconciliationParseResult(
                source_file_name=path.name,
                sheet_name=worksheet.title,
            )

            for excel_row_number, cells in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                result.total_excel_rows += 1

                if self._is_blank_row(cells):
                    result.skipped_blank_rows += 1
                    continue

                try:
                    parsed_values = self._parse_row(headers, cells)

                    ihx_ref_id = parsed_values.get("ihx_ref_id")
                    if not ihx_ref_id:
                        raise ValueError("IHX Ref Id is blank")

                    change_hash = self._calculate_change_hash(parsed_values)

                    result.parsed_rows.append(
                        ReconciliationRow(
                            source_row_number=excel_row_number,
                            values=parsed_values,
                            change_hash=change_hash,
                        )
                    )

                except Exception as exc:
                    result.row_errors.append(
                        ReconciliationRowError(
                            source_row_number=excel_row_number,
                            message=str(exc),
                        )
                    )

            logger.info(
                "[ReconciliationParser] Parsed file=%s sheet=%s "
                "excelRows=%s successful=%s blank=%s failed=%s",
                result.source_file_name,
                result.sheet_name,
                result.total_excel_rows,
                result.successful_rows,
                result.skipped_blank_rows,
                result.failed_rows,
            )

            return result

        finally:
            workbook.close()

    @staticmethod
    def _select_worksheet(workbook):
        if not workbook.sheetnames:
            raise ReconciliationParserError(
                "Reconciliation workbook contains no worksheets"
            )

        # Power BI currently exports the sheet as "Export".
        if "Export" in workbook.sheetnames:
            return workbook["Export"]

        # Safe fallback for future Power BI naming changes.
        return workbook[workbook.sheetnames[0]]

    @staticmethod
    def _read_headers(worksheet) -> List[str]:
        first_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not first_row:
            raise ReconciliationParserError(
                "Reconciliation workbook has no header row"
            )

        return [
            str(value).strip() if value is not None else ""
            for value in first_row
        ]

    @staticmethod
    def _validate_headers(actual_headers: List[str]) -> None:
        expected = list(EXPECTED_IHX_HEADERS)

        missing = [
            header
            for header in expected
            if header not in actual_headers
        ]

        unexpected = [
            header
            for header in actual_headers
            if header and header not in expected
        ]

        if missing or unexpected or len(actual_headers) != len(expected):
            messages = []

            if missing:
                messages.append(
                    "missing headers: " + ", ".join(missing)
                )

            if unexpected:
                messages.append(
                    "unexpected headers: " + ", ".join(unexpected)
                )

            if len(actual_headers) != len(expected):
                messages.append(
                    f"expected {len(expected)} columns but found "
                    f"{len(actual_headers)}"
                )

            raise ReconciliationParserError(
                "Invalid IHX reconciliation export structure; "
                + "; ".join(messages)
            )

    def _parse_row(
        self,
        headers: List[str],
        cells: tuple,
    ) -> Dict[str, Any]:
        parsed: Dict[str, Any] = {}

        for index, header in enumerate(headers):
            attribute_name = IHX_RECONCILIATION_FIELD_MAPPING[header]
            raw_value = cells[index] if index < len(cells) else None

            if attribute_name in DATE_FIELDS:
                value = self._normalize_date(raw_value)
            elif attribute_name in DECIMAL_FIELDS:
                value = self._normalize_decimal(raw_value)
            else:
                value = self._normalize_string(raw_value)

            parsed[attribute_name] = value

        return parsed

    @staticmethod
    def _is_blank_row(cells: tuple) -> bool:
        return all(
            value is None
            or (isinstance(value, str) and not value.strip())
            for value in cells
        )

    @staticmethod
    def _normalize_string(value: Any) -> Optional[str]:
        if value is None:
            return None

        if isinstance(value, bool):
            return "true" if value else "false"

        if isinstance(value, datetime):
            return value.isoformat()

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return format(value, "f").rstrip("0").rstrip(".")

        text = str(value).strip()

        if not text:
            return None

        if text.lower() in {
            "none",
            "null",
            "nan",
            "n/a",
            "na",
            "-",
        }:
            return None

        return text

    @staticmethod
    def _normalize_decimal(value: Any) -> Optional[Decimal]:
        if value is None:
            return None

        if isinstance(value, Decimal):
            return value.quantize(Decimal("0.01"))

        if isinstance(value, bool):
            raise ValueError(
                f"Boolean value is invalid for amount: {value}"
            )

        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal("0.01"))

        text = str(value).strip()

        if not text or text.lower() in {
            "none",
            "null",
            "nan",
            "n/a",
            "na",
            "-",
        }:
            return None

        cleaned = (
            text.replace(",", "")
            .replace("₹", "")
            .replace("INR", "")
            .strip()
        )

        # Support accounting notation: (1,250.00)
        is_negative = (
            cleaned.startswith("(")
            and cleaned.endswith(")")
        )
        if is_negative:
            cleaned = cleaned[1:-1].strip()

        try:
            amount = Decimal(cleaned)
        except InvalidOperation as exc:
            raise ValueError(
                f"Invalid amount value: {value!r}"
            ) from exc

        if is_negative:
            amount = -amount

        return amount.quantize(Decimal("0.01"))

    @staticmethod
    def _normalize_date(value: Any) -> Optional[date]:
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception as exc:
                raise ValueError(
                    f"Invalid Excel date value: {value!r}"
                ) from exc

        text = str(value).strip()

        if not text or text.lower() in {
            "none",
            "null",
            "nan",
            "n/a",
            "na",
            "-",
        }:
            return None

        supported_formats = (
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%b-%Y",
            "%d %b %Y",
            "%Y-%m-%d %H:%M:%S",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
        )

        for date_format in supported_formats:
            try:
                return datetime.strptime(
                    text,
                    date_format,
                ).date()
            except ValueError:
                continue

        try:
            return datetime.fromisoformat(text).date()
        except ValueError as exc:
            raise ValueError(
                f"Invalid date value: {value!r}"
            ) from exc

    @staticmethod
    def _calculate_change_hash(
        parsed_values: Dict[str, Any],
    ) -> str:
        serializable = {
            key: ReconciliationExcelParser._serialize_for_hash(value)
            for key, value in sorted(parsed_values.items())
        }

        canonical_json = json.dumps(
            serializable,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )

        return hashlib.sha256(
            canonical_json.encode("utf-8")
        ).hexdigest()

    @staticmethod
    def _serialize_for_hash(value: Any) -> Any:
        if isinstance(value, Decimal):
            return format(value, "f")

        if isinstance(value, (date, datetime)):
            return value.isoformat()

        return value
