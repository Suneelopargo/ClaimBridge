from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ReconciliationExcelReport:
    def generate(self, report_name, field_metadata, records):
        wb = Workbook()
        ws = wb.active
        ws.title = self._safe_sheet_name(report_name)

        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for idx, meta in enumerate(field_metadata, start=1):
            cell = ws.cell(row=1, column=idx, value=meta["displayName"])
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(int(meta.get("width", 120) / 7), 40))

        row_no = 2
        for record in records:
            for idx, meta in enumerate(field_metadata, start=1):
                cell = ws.cell(row=row_no, column=idx, value=record.get(meta["field"]))
                if meta.get("dataType") == "decimal":
                    cell.number_format = '#,##0.00'
                elif meta.get("dataType") == "date":
                    cell.number_format = 'dd-mmm-yyyy'
            row_no += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _safe_sheet_name(value):
        invalid = set(r'[]:*?/\\')
        cleaned = "".join("_" if ch in invalid else ch for ch in (value or "Reconciliation Report").strip())
        return cleaned[:31] or "Reconciliation Report"
