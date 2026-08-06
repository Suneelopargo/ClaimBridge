from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import CLAIM_PACKET_SEGREGATED_DIR


class PortfolioExcelService:
    def __init__(
        self,
        output_root: Path | None = None,
    ) -> None:
        self.output_root = (
            output_root
            or CLAIM_PACKET_SEGREGATED_DIR.parent
            / "validation-reports"
        )

    def generate_excel(
        self,
        portfolio: dict[str, Any] | None = None,
    ) -> Path:
        if portfolio is None:
            portfolio = self._load_portfolio_report()

        claim_reports = self._load_claim_reports(
            portfolio
        )

        workbook = Workbook()

        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self._build_executive_summary_sheet(
            workbook,
            portfolio,
        )
        self._build_claim_summary_sheet(
            workbook,
            portfolio,
        )
        self._build_document_matrix_sheet(
            workbook,
            claim_reports,
        )
        self._build_review_queue_sheet(
            workbook,
            claim_reports,
        )
        self._build_failure_queue_sheet(
            workbook,
            claim_reports,
        )
        self._build_rule_catalogue_sheet(
            workbook,
            claim_reports,
            portfolio,
        )

        portfolio_directory = (
            self.output_root
            / "portfolio"
        )

        portfolio_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        output_path = (
            portfolio_directory
            / "HCG_Claim_Validation_Report.xlsx"
        )

        workbook.save(output_path)

        return output_path

    def _load_portfolio_report(
        self,
    ) -> dict[str, Any]:
        path = (
            self.output_root
            / "portfolio"
            / "portfolio_validation_report.json"
        )

        if not path.exists():
            raise FileNotFoundError(
                "Portfolio validation report "
                "has not been generated"
            )

        with path.open(
            "r",
            encoding="utf-8",
        ) as file:
            data = json.load(file)

        if not isinstance(data, dict):
            raise ValueError(
                "Portfolio report must be "
                "a JSON object"
            )

        return data

    def _load_claim_reports(
        self,
        portfolio: dict[str, Any],
    ) -> list[dict[str, Any]]:
        reports: list[
            dict[str, Any]
        ] = []

        for claim in (
            portfolio.get("claims")
            or []
        ):
            claim_id = str(
                claim.get("claimId")
                or ""
            ).strip()

            if not claim_id:
                continue

            path = (
                self.output_root
                / "claims"
                / claim_id
                / "document_validation_report.json"
            )

            if not path.exists():
                continue

            with path.open(
                "r",
                encoding="utf-8",
            ) as file:
                report = json.load(file)

            if isinstance(report, dict):
                reports.append(report)

        return reports

    def _build_executive_summary_sheet(
        self,
        workbook: Workbook,
        portfolio: dict[str, Any],
    ) -> None:
        sheet = workbook.create_sheet(
            "Executive Summary"
        )

        summary = (
            portfolio.get("summary")
            or {}
        )

        sheet["A1"] = (
            "HCG Claim Validation "
            "Portfolio Report"
        )

        sheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        rows = [
            (
                "Generated At",
                portfolio.get(
                    "generatedAt"
                ),
            ),
            (
                "Total Claims",
                summary.get(
                    "successfulClaims",
                    0,
                ),
            ),
            (
                "Total Documents",
                summary.get(
                    "totalDocuments",
                    0,
                ),
            ),
            (
                "Applicable Checks",
                summary.get(
                    "applicableChecks",
                    0,
                ),
            ),
            (
                "Pass Count",
                summary.get(
                    "passCount",
                    0,
                ),
            ),
            (
                "Fail Count",
                summary.get(
                    "failCount",
                    0,
                ),
            ),
            (
                "Review Count",
                summary.get(
                    "reviewCount",
                    0,
                ),
            ),
            (
                "Not Applicable Count",
                summary.get(
                    "naCount",
                    0,
                ),
            ),
            (
                "Confirmed Readiness %",
                summary.get(
                    "confirmedReadinessPercent",
                    0,
                ),
            ),
            (
                "Decision Coverage %",
                summary.get(
                    "decisionCoveragePercent",
                    0,
                ),
            ),
            (
                "Human Review Burden %",
                summary.get(
                    "humanReviewBurdenPercent",
                    0,
                ),
            ),
        ]

        start_row = 3

        for index, (
            label,
            value,
        ) in enumerate(
            rows,
            start=start_row,
        ):
            sheet.cell(
                row=index,
                column=1,
                value=label,
            )

            sheet.cell(
                row=index,
                column=2,
                value=value,
            )

        sheet["A16"] = (
            "Important Scope Note"
        )
        sheet["A16"].font = Font(
            bold=True
        )

        sheet["A17"] = (
            "This MVP report currently "
            "executes 28 configured checks "
            "mapped against the broader "
            "72-rule HCG catalogue."
        )

        sheet.merge_cells(
            "A17:D18"
        )

        sheet["A17"].alignment = (
            Alignment(
                wrap_text=True,
                vertical="top",
            )
        )

        sheet.column_dimensions[
            "A"
        ].width = 32

        sheet.column_dimensions[
            "B"
        ].width = 22

    def _build_claim_summary_sheet(
        self,
        workbook: Workbook,
        portfolio: dict[str, Any],
    ) -> None:
        sheet = workbook.create_sheet(
            "Claim Summary"
        )

        headers = [
            "Claim ID",
            "Patient",
            "Payer",
            "Documents",
            "Pass",
            "Fail",
            "Review",
            "NA",
            "Confirmed Readiness %",
            "Decision Coverage %",
            "Review Burden %",
            "Overall Status",
        ]

        self._write_headers(
            sheet,
            headers,
        )

        for row_index, claim in enumerate(
            portfolio.get("claims")
            or [],
            start=2,
        ):
            values = [
                claim.get("claimId"),
                claim.get("patientName"),
                claim.get("payerCode"),
                claim.get(
                    "documentCount",
                    0,
                ),
                claim.get(
                    "passCount",
                    0,
                ),
                claim.get(
                    "failCount",
                    0,
                ),
                claim.get(
                    "reviewCount",
                    0,
                ),
                claim.get(
                    "naCount",
                    0,
                ),
                claim.get(
                    "confirmedReadinessPercent",
                    0,
                ),
                claim.get(
                    "decisionCoveragePercent",
                    0,
                ),
                claim.get(
                    "humanReviewBurdenPercent",
                    0,
                ),
                claim.get(
                    "overallStatus"
                ),
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

        self._autosize(sheet)

    def _build_document_matrix_sheet(
        self,
        workbook: Workbook,
        claim_reports: list[
            dict[str, Any]
        ],
    ) -> None:
        sheet = workbook.create_sheet(
            "Document Matrix"
        )

        rule_catalogue = (
            self._collect_rule_catalogue(
                claim_reports
            )
        )

        headers = [
            "Claim ID",
            "Patient",
            "Document ID",
            "Document Type",
            "Display Name",
            "Pages",
        ] + [
            str(rule["ruleNumber"])
            for rule in rule_catalogue
        ]

        self._write_headers(
            sheet,
            headers,
        )

        for column_index, rule in enumerate(
            rule_catalogue,
            start=7,
        ):
            cell = sheet.cell(
                row=1,
                column=column_index,
            )

            cell.comment = Comment(
                (
                    f'Rule {rule["ruleNumber"]}: '
                    f'{rule["ruleName"]}'
                ),
                "ClaimBridge",
            )

        row_index = 2

        for report in claim_reports:
            claim_id = report.get(
                "claimId"
            )

            patient_name = report.get(
                "patientName"
            )

            for document in (
                report.get("documents")
                or []
            ):
                result_lookup = {
                    int(
                        result.get(
                            "ruleNumber"
                        )
                    ): result.get(
                        "status"
                    )
                    for result in (
                        document.get(
                            "ruleResults"
                        )
                        or []
                    )
                    if result.get(
                        "ruleNumber"
                    )
                    is not None
                }

                values = [
                    claim_id,
                    patient_name,
                    document.get(
                        "documentId"
                    ),
                    document.get(
                        "documentType"
                    ),
                    document.get(
                        "displayName"
                    ),
                    ", ".join(
                        str(page)
                        for page in (
                            document.get(
                                "pageNumbers"
                            )
                            or []
                        )
                    ),
                ]

                for rule in rule_catalogue:
                    values.append(
                        result_lookup.get(
                            int(
                                rule[
                                    "ruleNumber"
                                ]
                            ),
                            "NA",
                        )
                    )

                for column_index, value in enumerate(
                    values,
                    start=1,
                ):
                    sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value,
                    )

                row_index += 1

        sheet.freeze_panes = "G2"
        sheet.auto_filter.ref = (
            sheet.dimensions
        )

        self._autosize(
            sheet,
            max_width=24,
        )

    def _build_review_queue_sheet(
        self,
        workbook: Workbook,
        claim_reports: list[
            dict[str, Any]
        ],
    ) -> None:
        sheet = workbook.create_sheet(
            "Review Queue"
        )

        headers = [
            "Claim ID",
            "Patient",
            "Document",
            "Document Type",
            "Pages",
            "Rule No.",
            "Rule Name",
            "Reason",
            "Confidence",
            "Expected Value",
            "Actual Value",
        ]

        self._write_headers(
            sheet,
            headers,
        )

        row_index = 2

        for report in claim_reports:
            insights = (
                report.get("insights")
                or {}
            )

            for item in (
                insights.get(
                    "priorityReviewQueue"
                )
                or []
            ):
                evidence = (
                    item.get("evidence")
                    or {}
                )

                values = [
                    report.get(
                        "claimId"
                    ),
                    report.get(
                        "patientName"
                    ),
                    item.get(
                        "displayName"
                    ),
                    item.get(
                        "documentType"
                    ),
                    ", ".join(
                        str(page)
                        for page in (
                            item.get(
                                "pageNumbers"
                            )
                            or []
                        )
                    ),
                    item.get(
                        "ruleNumber"
                    ),
                    item.get(
                        "ruleName"
                    ),
                    item.get(
                        "reason"
                    ),
                    item.get(
                        "confidence"
                    ),
                    self._extract_expected_value(
                        evidence
                    ),
                    self._extract_actual_value(
                        evidence
                    ),
                ]

                for column_index, value in enumerate(
                    values,
                    start=1,
                ):
                    sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value,
                    )

                row_index += 1

        self._autosize(
            sheet,
            max_width=45,
        )

    def _build_failure_queue_sheet(
        self,
        workbook: Workbook,
        claim_reports: list[
            dict[str, Any]
        ],
    ) -> None:
        sheet = workbook.create_sheet(
            "Failure Queue"
        )

        headers = [
            "Claim ID",
            "Patient",
            "Document",
            "Document Type",
            "Pages",
            "Rule No.",
            "Rule Name",
            "Reason",
            "Confidence",
            "Expected Value",
            "Actual Value",
        ]

        self._write_headers(
            sheet,
            headers,
        )

        row_index = 2

        for report in claim_reports:
            insights = (
                report.get("insights")
                or {}
            )

            for item in (
                insights.get(
                    "failedItems"
                )
                or []
            ):
                evidence = (
                    item.get("evidence")
                    or {}
                )

                values = [
                    report.get(
                        "claimId"
                    ),
                    report.get(
                        "patientName"
                    ),
                    item.get(
                        "displayName"
                    ),
                    item.get(
                        "documentType"
                    ),
                    ", ".join(
                        str(page)
                        for page in (
                            item.get(
                                "pageNumbers"
                            )
                            or []
                        )
                    ),
                    item.get(
                        "ruleNumber"
                    ),
                    item.get(
                        "ruleName"
                    ),
                    item.get(
                        "reason"
                    ),
                    item.get(
                        "confidence"
                    ),
                    self._extract_expected_value(
                        evidence
                    ),
                    self._extract_actual_value(
                        evidence
                    ),
                ]

                for column_index, value in enumerate(
                    values,
                    start=1,
                ):
                    sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value,
                    )

                row_index += 1

        self._autosize(
            sheet,
            max_width=45,
        )

    def _build_rule_catalogue_sheet(
        self,
        workbook: Workbook,
        claim_reports: list[
            dict[str, Any]
        ],
        portfolio: dict[str, Any],
    ) -> None:
        sheet = workbook.create_sheet(
            "Rule Catalogue"
        )

        headers = [
            "Rule Number",
            "Rule Code",
            "Rule Name",
            "Execution Status",
        ]

        self._write_headers(
            sheet,
            headers,
        )

        configured_rules = (
            self._collect_rule_catalogue(
                claim_reports
            )
        )

        configured_numbers = {
            int(rule["ruleNumber"])
            for rule in configured_rules
        }

        row_index = 2

        for rule in configured_rules:
            values = [
                rule.get(
                    "ruleNumber"
                ),
                rule.get(
                    "ruleCode"
                ),
                rule.get(
                    "ruleName"
                ),
                "CONFIGURED",
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

            row_index += 1

        for rule_number in range(1, 73):
            if rule_number in (
                configured_numbers
            ):
                continue

            values = [
                rule_number,
                (
                    f"HCG-VAL-"
                    f"{rule_number:03d}"
                ),
                (
                    f"Rule {rule_number}"
                ),
                "NOT YET CONFIGURED",
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

            row_index += 1

        self._autosize(sheet)

    @staticmethod
    def _collect_rule_catalogue(
        claim_reports: list[
            dict[str, Any]
        ],
    ) -> list[dict[str, Any]]:
        lookup: dict[
            int,
            dict[str, Any],
        ] = {}

        for report in claim_reports:
            for rule in (
                report.get("rules")
                or []
            ):
                try:
                    rule_number = int(
                        rule.get(
                            "ruleNumber"
                        )
                    )
                except (
                    TypeError,
                    ValueError,
                ):
                    continue

                lookup[rule_number] = rule

        return [
            lookup[key]
            for key in sorted(lookup)
        ]

    @staticmethod
    def _extract_expected_value(
        evidence: dict[str, Any],
    ) -> Any:
        for key in (
            "expectedPatientName",
            "expectedValue",
            "expected",
        ):
            if key in evidence:
                return evidence[key]

        return None

    @staticmethod
    def _extract_actual_value(
        evidence: dict[str, Any],
    ) -> Any:
        for key in (
            "actualPatientName",
            "actualValue",
            "actual",
            "documentDate",
        ):
            if key in evidence:
                return evidence[key]

        return None

    @staticmethod
    def _write_headers(
        sheet,
        headers: list[str],
    ) -> None:
        fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True
            )
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    @staticmethod
    def _autosize(
        sheet,
        max_width: int = 35,
    ) -> None:
        for column_cells in (
            sheet.columns
        ):
            max_length = 0

            column_index = (
                column_cells[0]
                .column
            )

            for cell in column_cells:
                if cell.value is None:
                    continue

                text = str(
                    cell.value
                )

                max_length = max(
                    max_length,
                    len(text),
                )

            width = min(
                max_length + 2,
                max_width,
            )

            sheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = max(
                width,
                10,
            )