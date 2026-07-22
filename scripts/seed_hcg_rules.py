from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session


# Project root:
# C:\Users\Projects\india-claims-automation
PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


# Import application modules only AFTER adding PROJECT_ROOT
import app.models  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models.validation_rule import ValidationRule  # noqa: E402
from app.models.validation_rule_document_type import (  # noqa: E402
    ValidationRuleDocumentType,
)
from app.models.validation_rule_version import (  # noqa: E402
    ValidationRuleVersion,
)


DEFAULT_WORKBOOK = (
    PROJECT_ROOT
    / "data"
    / "rulebooks"
    / "Rule Book For RPA Sulution.xlsx"
)

SHEET_NAME = "RuleBook"
IMPORTED_BY = "HCG_RULEBOOK_IMPORT"


# ---------------------------------------------------------------------------
# Workbook columns
# ---------------------------------------------------------------------------

PAYER_COLUMNS = {
    "Insurance(TPA)": "INSURANCE_TPA",
    "Govt Schemes": "GOVT_SCHEME",
    "Corporate PSU": "CORPORATE_PSU",
    "CGHS": "CGHS",
    "ECHS": "ECHS",
    "ESIC": "ESIC",
}


# ---------------------------------------------------------------------------
# Rule classification
# ---------------------------------------------------------------------------

CATEGORY_RULES: list[tuple[tuple[str, ...], str]] = [
    (
        (
            "document",
            "report completeness",
            "file format",
            "segregation",
            "naming",
            "missing report",
        ),
        "DOCUMENT",
    ),
    (
        (
            "patient identity",
            "kyc",
            "claim form",
            "signature",
        ),
        "IDENTITY",
    ),
    (
        (
            "admission",
            "discharge",
            "date consistency",
            "chronology",
            "death",
            "transfer",
        ),
        "CHRONOLOGY",
    ),
    (
        (
            "diagnosis",
            "procedure",
            "investigation",
            "radiology",
            "laboratory",
            "blood",
            "pharmacy",
            "implant",
            "clinical",
            "consent",
            "mlc",
        ),
        "CLINICAL",
    ),
    (
        (
            "billing",
            "invoice",
            "tariff",
            "room rent",
            "policy limit",
            "non-payable",
            "package",
            "consumable",
            "deduction",
            "amount",
        ),
        "FINANCIAL",
    ),
    (
        (
            "coding",
            "icd",
            "cpt",
        ),
        "CODING",
    ),
    (
        (
            "duplicate",
        ),
        "DUPLICATE",
    ),
    (
        (
            "portal",
            "query",
            "submission",
        ),
        "PORTAL",
    ),
    (
        (
            "audit",
            "version control",
        ),
        "AUDIT",
    ),
    (
        (
            "analytics",
            "readiness",
            "exception",
        ),
        "ANALYTICS",
    ),
]


RULE_TYPE_RULES: list[tuple[tuple[str, ...], str]] = [
    (
        (
            "mandatory document",
            "document availability",
            "checklist",
        ),
        "DOCUMENT_PRESENCE",
    ),
    (
        (
            "claim form",
            "field",
            "signature",
            "report completeness",
            "document quality",
            "document naming",
            "file format",
        ),
        "FIELD_REQUIRED",
    ),
    (
        (
            "identity",
            "match",
            "consistency",
            "validation",
        ),
        "CROSS_DOCUMENT_MATCH",
    ),
    (
        (
            "admission",
            "discharge",
            "date consistency",
            "chronology",
        ),
        "DATE_SEQUENCE",
    ),
    (
        (
            "amount",
            "billing",
            "invoice",
            "tariff",
            "room rent",
            "policy limit",
            "non-payable",
            "package",
            "deduction",
        ),
        "AMOUNT_RECONCILIATION",
    ),
    (
        (
            "duplicate document",
            "duplicate billing",
        ),
        "DUPLICATE_CHECK",
    ),
    (
        (
            "clinical",
            "diagnosis",
            "procedure",
            "investigation",
            "radiology",
            "laboratory",
            "pharmacy",
            "implant",
            "consent",
            "mlc",
        ),
        "CLINICAL_CONSISTENCY",
    ),
    (
        (
            "portal",
            "submission",
            "query",
        ),
        "PORTAL_VALIDATION",
    ),
]


SEVERITY_RULES: list[tuple[tuple[str, ...], str]] = [
    (
        (
            "mandatory document",
            "patient identity",
            "admission",
            "discharge",
            "approval",
            "final bill",
            "portal submission",
        ),
        "CRITICAL",
    ),
    (
        (
            "diagnosis",
            "procedure",
            "billing",
            "policy limit",
            "tariff",
            "duplicate",
        ),
        "ERROR",
    ),
    (
        (
            "quality",
            "naming",
            "format",
            "analytics",
            "feedback",
        ),
        "WARNING",
    ),
]


# ---------------------------------------------------------------------------
# Detailed checklist document taxonomy
# ---------------------------------------------------------------------------

CHECKLIST_DOCUMENT_TYPES: dict[int, list[str]] = {
    51: ["PREAUTHORIZATION_FORM"],
    52: ["CLAIM_FORM"],
    53: ["GIPSA_DECLARATION"],
    54: [
        "KYC_DOCUMENT",
        "PROPOSER_ID_PROOF",
    ],
    55: ["PATIENT_ID_PROOF"],
    56: [
        "PRESCRIPTION",
        "INVESTIGATION_REPORT",
        "RADIOLOGY_REPORT",
    ],
    57: ["CONSENT_FORM"],
    58: [
        "APPROVAL_LETTER",
        "GOP_PRE_APPROVAL",
        "GOP_FINAL_APPROVAL",
        "CASHLESS_AUTHORIZATION_LETTER",
    ],
    59: ["DISCHARGE_SUMMARY"],
    60: [
        "CASE_PAPER",
        "OT_NOTES",
    ],
    61: [
        "INVESTIGATION_REPORT",
        "LAB_REPORT",
    ],
    62: [
        "LAB_REPORT",
        "RADIOLOGY_REPORT",
        "INVESTIGATION_REPORT",
    ],
    63: [
        "RADIOLOGY_REPORT",
        "RADIOLOGY_FILM",
    ],
    64: [
        "IMPLANT_STICKER_INVOICE",
        "CHEMOTHERAPY_WRAPPER",
        "VENDOR_INVOICE",
    ],
    65: ["BLOOD_COMPONENT_STICKER"],
    66: ["PATIENT_PHOTO"],
    67: [
        "MLC_DOCUMENT",
        "POLICE_FIR",
    ],
    68: [
        "FINAL_HOSPITAL_BILL",
        "DETAILED_BILL_BREAKUP",
        "BILL_CONTINUATION",
    ],
    69: [
        "PAYMENT_RECEIPT",
        "REFUND_RECEIPT",
    ],
    70: [
        "PACKAGE_BREAKUP",
        "DETAILED_BILL_BREAKUP",
    ],
    71: [
        "PHARMACY_DETAILS",
        "PHARMACY_BILL",
    ],
    72: ["FEEDBACK_FORM"],
}


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_multiline_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    return "\n".join(lines)


def slugify(value: str) -> str:
    value = value.upper().strip()
    value = re.sub(r"[^A-Z0-9]+", "_", value)
    return value.strip("_")


def is_applicable(value: Any) -> bool:
    """
    Treat only an explicit 'Applicable' value as applicable.

    The workbook contains at least one malformed cell, such as:
    'App+K40+I49'. It is intentionally not treated as applicable and is
    preserved under raw workbook values for manual review.
    """

    return normalize_text(value).lower() == "applicable"


def derive_category(name: str, logic: str) -> str:
    combined = f"{name} {logic}".lower()

    for keywords, category in CATEGORY_RULES:
        if any(keyword in combined for keyword in keywords):
            return category

    return "GENERAL"


def derive_rule_type(
    serial_number: int,
    name: str,
    logic: str,
) -> str:
    if serial_number >= 51:
        return "DOCUMENT_PRESENCE"

    combined = f"{name} {logic}".lower()

    # More specific checks should be evaluated before generic "validation".
    if "duplicate document" in combined or "duplicate billing" in combined:
        return "DUPLICATE_CHECK"

    if any(
        keyword in combined
        for keyword in (
            "admission",
            "discharge",
            "date consistency",
            "chronology",
        )
    ):
        return "DATE_SEQUENCE"

    if any(
        keyword in combined
        for keyword in (
            "billing",
            "invoice",
            "tariff",
            "room rent",
            "policy limit",
            "non-payable",
            "package",
            "deduction",
        )
    ):
        return "AMOUNT_RECONCILIATION"

    if any(
        keyword in combined
        for keyword in (
            "mandatory document",
            "document availability",
            "checklist",
        )
    ):
        return "DOCUMENT_PRESENCE"

    if any(
        keyword in combined
        for keyword in (
            "diagnosis",
            "procedure",
            "investigation",
            "radiology",
            "laboratory",
            "pharmacy",
            "implant",
            "clinical",
            "consent",
            "mlc",
        )
    ):
        return "CLINICAL_CONSISTENCY"

    if any(
        keyword in combined
        for keyword in (
            "portal",
            "submission",
            "query",
        )
    ):
        return "PORTAL_VALIDATION"

    if any(
        keyword in combined
        for keyword in (
            "match",
            "consistency",
            "identity",
        )
    ):
        return "CROSS_DOCUMENT_MATCH"

    if any(
        keyword in combined
        for keyword in (
            "field",
            "signature",
            "quality",
            "format",
            "naming",
            "completeness",
        )
    ):
        return "FIELD_REQUIRED"

    return "MANUAL_REVIEW"


def derive_severity(name: str, logic: str) -> str:
    combined = f"{name} {logic}".lower()

    for keywords, severity in SEVERITY_RULES:
        if any(keyword in combined for keyword in keywords):
            return severity

    return "WARNING"


def extract_page_expectation(value: Any) -> str | None:
    text = normalize_text(value)

    return text or None


def build_payer_applicability(
    headers: list[str],
    row: list[Any],
) -> dict[str, Any]:
    payer_matrix: dict[str, bool] = {}
    raw_matrix: dict[str, str] = {}

    header_index = {
        normalize_text(header): index
        for index, header in enumerate(headers)
    }

    for workbook_column, payer_code in PAYER_COLUMNS.items():
        column_index = header_index.get(workbook_column)

        if column_index is None:
            continue

        raw_value = (
            row[column_index]
            if column_index < len(row)
            else None
        )

        payer_matrix[payer_code] = is_applicable(raw_value)
        raw_matrix[payer_code] = normalize_text(raw_value)

    return {
        "operator": "PAYER_IN",
        "payerTypes": [
            payer_code
            for payer_code, applicable in payer_matrix.items()
            if applicable
        ],
        "payerMatrix": payer_matrix,
        "rawPayerValues": raw_matrix,
    }


def build_validation_expression(
    serial_number: int,
    rule_type: str,
    logic: str,
    applicable_payer_type: str,
) -> dict[str, Any]:
    if serial_number >= 51:
        return {
            "operator": "DOCUMENT_EXISTS",
            "documentTypes": CHECKLIST_DOCUMENT_TYPES.get(
                serial_number,
                [],
            ),
            "requirementText": logic,
            "expectedPageCount": (
                applicable_payer_type or None
            ),
            "executionMode": "DRAFT",
        }

    return {
        "operator": rule_type,
        "businessLogic": logic,
        "executionMode": "MANUAL_REVIEW",
        "implementationStatus": "NOT_IMPLEMENTED",
    }


def build_failure_message(rule_name: str) -> str:
    return f"{rule_name} validation failed or requires review."


def build_success_message(rule_name: str) -> str:
    return f"{rule_name} validation completed successfully."


def get_or_create_rule(
    db: Session,
    *,
    rule_code: str,
    rule_name: str,
    description: str,
    category: str,
    rule_type: str,
    severity: str,
) -> ValidationRule:
    rule = (
        db.query(ValidationRule)
        .filter(ValidationRule.rule_code == rule_code)
        .one_or_none()
    )

    if rule is None:
        rule = ValidationRule(
            rule_code=rule_code,
            rule_name=rule_name,
            description=description,
            category=category,
            rule_type=rule_type,
            payer_code=None,
            severity=severity,
            is_active=True,
            created_by=IMPORTED_BY,
            updated_by=IMPORTED_BY,
        )

        db.add(rule)
        db.flush()

        return rule

    # Synchronize only editable master data.
    rule.rule_name = rule_name
    rule.description = description
    rule.category = category
    rule.rule_type = rule_type
    rule.severity = severity
    rule.is_active = True
    rule.updated_by = IMPORTED_BY

    db.flush()

    return rule


def upsert_draft_version(
    db: Session,
    *,
    rule: ValidationRule,
    applicability_expression: dict[str, Any],
    validation_expression: dict[str, Any],
) -> ValidationRuleVersion:
    """
    Import into version 1 while it remains unpublished.

    Once version 1 is published, the importer creates or updates the next
    unpublished draft version instead of altering the published definition.
    """

    versions = (
        db.query(ValidationRuleVersion)
        .filter(ValidationRuleVersion.rule_id == rule.id)
        .order_by(ValidationRuleVersion.version_number.asc())
        .all()
    )

    draft_version = next(
        (
            version
            for version in versions
            if not version.is_published
        ),
        None,
    )

    if draft_version is None:
        next_version_number = (
            max(
                (
                    version.version_number
                    for version in versions
                ),
                default=0,
            )
            + 1
        )

        draft_version = ValidationRuleVersion(
            rule_id=rule.id,
            version_number=next_version_number,
            effective_from=date.today(),
            is_published=False,
            created_by=IMPORTED_BY,
        )

        db.add(draft_version)

    draft_version.applicability_expression = (
        applicability_expression
    )
    draft_version.validation_expression = (
        validation_expression
    )
    draft_version.success_message = build_success_message(
        rule.rule_name
    )
    draft_version.failure_message = build_failure_message(
        rule.rule_name
    )
    draft_version.change_reason = (
        "Imported from HCG Rule Book workbook"
    )

    db.flush()

    return draft_version


def replace_document_mappings(
    db: Session,
    *,
    rule: ValidationRule,
    document_types: list[str],
) -> None:
    (
        db.query(ValidationRuleDocumentType)
        .filter(
            ValidationRuleDocumentType.rule_id == rule.id
        )
        .delete(synchronize_session=False)
    )

    for document_type in sorted(set(document_types)):
        mapping = ValidationRuleDocumentType(
            rule_id=rule.id,
            document_type=document_type,
            document_role="REQUIRED",
            is_mandatory=True,
        )

        db.add(mapping)


# ---------------------------------------------------------------------------
# Workbook import
# ---------------------------------------------------------------------------

def seed_hcg_rules(
    workbook_path: Path,
    *,
    dry_run: bool = False,
) -> dict[str, int]:
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"HCG rule book not found: {workbook_path}"
        )

    workbook = load_workbook(
        filename=workbook_path,
        data_only=True,
        read_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet '{SHEET_NAME}' was not found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    sheet = workbook[SHEET_NAME]

    # The workbook data begins in column B and ends in column K.
    rows = list(
        sheet.iter_rows(
            min_col=2,
            max_col=11,
            values_only=True,
        )
    )

    if not rows:
        raise ValueError("The HCG workbook contains no rows")

    headers = [
        normalize_text(value)
        for value in rows[0]
    ]

    expected_headers = {
        "Sl No",
        "Validation Point",
        "AI/RPA Validation Logic",
    }

    if not expected_headers.issubset(set(headers)):
        raise ValueError(
            "Unexpected HCG rule-book format. "
            f"Headers found: {headers}"
        )

    header_index = {
        header: index
        for index, header in enumerate(headers)
    }

    stats = {
        "rowsRead": 0,
        "rulesCreatedOrUpdated": 0,
        "businessRules": 0,
        "checklistRules": 0,
        "documentMappings": 0,
        "skippedRows": 0,
    }

    db = SessionLocal()

    try:
        for row in rows[1:]:
            raw_serial_number = row[
                header_index["Sl No"]
            ]

            try:
                serial_number = int(raw_serial_number)
            except (TypeError, ValueError):
                stats["skippedRows"] += 1
                continue

            if serial_number < 1 or serial_number > 72:
                stats["skippedRows"] += 1
                continue

            stats["rowsRead"] += 1

            rule_name = normalize_text(
                row[header_index["Validation Point"]]
            )

            logic = normalize_multiline_text(
                row[
                    header_index[
                        "AI/RPA Validation Logic"
                    ]
                ]
            )

            applicable_payer_type = normalize_text(
                row[
                    header_index[
                        "Applicable Payor Type"
                    ]
                ]
            )

            if not rule_name:
                stats["skippedRows"] += 1
                continue

            category = derive_category(rule_name, logic)
            rule_type = derive_rule_type(
                serial_number,
                rule_name,
                logic,
            )
            severity = derive_severity(rule_name, logic)

            rule_prefix = (
                "HCG-BUS"
                if serial_number <= 50
                else "HCG-DOC"
            )

            rule_code = (
                f"{rule_prefix}-{serial_number:03d}"
            )

            description = (
                f"HCG rule book item {serial_number}: "
                f"{rule_name}"
            )

            applicability_expression = (
                build_payer_applicability(
                    headers,
                    list(row),
                )
            )

            applicability_expression.update(
                {
                    "workbookRuleNumber": serial_number,
                    "applicablePayorTypeText": (
                        applicable_payer_type
                    ),
                    "source": "HCG_RULE_BOOK",
                }
            )

            validation_expression = (
                build_validation_expression(
                    serial_number,
                    rule_type,
                    logic,
                    applicable_payer_type,
                )
            )

            rule = get_or_create_rule(
                db,
                rule_code=rule_code,
                rule_name=rule_name,
                description=description,
                category=category,
                rule_type=rule_type,
                severity=severity,
            )

            upsert_draft_version(
                db,
                rule=rule,
                applicability_expression=(
                    applicability_expression
                ),
                validation_expression=(
                    validation_expression
                ),
            )

            document_types = (
                CHECKLIST_DOCUMENT_TYPES.get(
                    serial_number,
                    [],
                )
            )

            if serial_number >= 51:
                replace_document_mappings(
                    db,
                    rule=rule,
                    document_types=document_types,
                )

                stats["checklistRules"] += 1
                stats["documentMappings"] += len(
                    set(document_types)
                )
            else:
                stats["businessRules"] += 1

            stats["rulesCreatedOrUpdated"] += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

        return stats

    except Exception:
        db.rollback()
        raise

    finally:
        db.close()
        workbook.close()


# ---------------------------------------------------------------------------
# Command-line entry point
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Import the HCG Excel rule book into "
            "India Claims Automation PostgreSQL tables."
        )
    )

    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=(
            "Path to the HCG Rule Book Excel file. "
            f"Default: {DEFAULT_WORKBOOK}"
        ),
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help=(
            "Parse and validate the workbook, then roll back "
            "all database changes."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    stats = seed_hcg_rules(
        workbook_path=args.file.resolve(),
        dry_run=args.dry_run,
    )

    action = (
        "Dry run completed"
        if args.dry_run
        else "HCG rules seeded successfully"
    )

    print(action)

    for key, value in stats.items():
        print(f"  {key}: {value}")


if __name__ == "__main__":
    main()